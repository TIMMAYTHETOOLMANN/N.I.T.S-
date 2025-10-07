"""
Excel File Processor
Handles both binary XLS (BIFF8) and OpenXML XLSX formats
"""

import pandas as pd
import os
from typing import Dict, Any, Optional, Union

class ExcelProcessor:
    """Unified Excel processor supporting .xls and .xlsx formats"""
    
    @staticmethod
    def detect_excel_format(file_path: str) -> str:
        """Detect Excel format by reading file signature"""
        
        with open(file_path, 'rb') as f:
            signature = f.read(8)
        
        # XLSX/XLSM are ZIP archives
        if signature.startswith(b'PK\x03\x04'):
            return 'xlsx'
        
        # XLS has OLE/CFB signature
        if signature.startswith(b'\xD0\xCF\x11\xE0\xA1\xB1\x1A\xE1'):
            return 'xls'
        
        # Fallback to extension
        ext = os.path.splitext(file_path)[1].lower()
        if ext in ['.xls', '.xlsx', '.xlsm']:
            return ext[1:]  # Remove dot
        
        raise ValueError(f"Unknown Excel format for {file_path}")
    
    @staticmethod
    def read_excel_auto(
        file_path: str,
        sheet_name: Optional[Union[str, int]] = 0
    ) -> pd.DataFrame:
        """
        Automatically detect format and use correct engine
        
        Args:
            file_path: Path to Excel file
            sheet_name: Sheet name or index to read (default: first sheet)
        
        Returns:
            pandas DataFrame with Excel data
        """
        
        format_type = ExcelProcessor.detect_excel_format(file_path)
        
        print(f"Detected Excel format: {format_type}")
        
        if format_type == 'xls':
            # Binary XLS format - REQUIRES xlrd
            try:
                df = pd.read_excel(
                    file_path,
                    sheet_name=sheet_name,
                    engine='xlrd'
                )
                print(f"✓ Read XLS file with xlrd: {len(df)} rows")
            except ImportError:
                raise ImportError(
                    "xlrd library required for .xls files. "
                    "Install with: pip install xlrd==2.0.1"
                )
        
        elif format_type in ['xlsx', 'xlsm']:
            # OpenXML format - REQUIRES openpyxl
            try:
                df = pd.read_excel(
                    file_path,
                    sheet_name=sheet_name,
                    engine='openpyxl'
                )
                print(f"✓ Read XLSX file with openpyxl: {len(df)} rows")
            except ImportError:
                raise ImportError(
                    "openpyxl library required for .xlsx files. "
                    "Install with: pip install openpyxl"
                )
        
        else:
            raise ValueError(f"Unsupported Excel format: {format_type}")
        
        # Validate data was read
        if df.empty:
            raise ValueError(f"Excel file appears empty: {file_path}")
        
        return df
    
    @staticmethod
    def get_excel_metadata(file_path: str) -> Dict[str, Any]:
        """Get Excel file metadata without reading all data"""
        
        format_type = ExcelProcessor.detect_excel_format(file_path)
        
        metadata = {
            'file_path': file_path,
            'file_size': os.path.getsize(file_path),
            'format': format_type,
            'sheets': []
        }
        
        if format_type == 'xls':
            import xlrd
            workbook = xlrd.open_workbook(file_path)
            metadata['sheets'] = workbook.sheet_names()
            metadata['engine'] = 'xlrd'
        
        elif format_type in ['xlsx', 'xlsm']:
            from openpyxl import load_workbook
            workbook = load_workbook(file_path, read_only=True)
            metadata['sheets'] = workbook.sheetnames
            metadata['engine'] = 'openpyxl'
            workbook.close()
        
        return metadata
    
    @staticmethod
    def read_all_sheets(file_path: str) -> Dict[str, pd.DataFrame]:
        """Read all sheets from Excel file"""
        
        format_type = ExcelProcessor.detect_excel_format(file_path)
        
        engine = 'xlrd' if format_type == 'xls' else 'openpyxl'
        
        # Read all sheets
        all_sheets = pd.read_excel(
            file_path,
            sheet_name=None,  # None means read all sheets
            engine=engine
        )
        
        print(f"✓ Read {len(all_sheets)} sheets from {format_type} file")
        
        return all_sheets


# Test function
if __name__ == "__main__":
    import sys
    
    if len(sys.argv) > 1:
        test_file = sys.argv[1]
        
        print(f"\nTesting Excel processor with: {test_file}")
        
        # Get metadata
        metadata = ExcelProcessor.get_excel_metadata(test_file)
        print(f"\nMetadata: {metadata}")
        
        # Read first sheet
        df = ExcelProcessor.read_excel_auto(test_file)
        print(f"\nFirst sheet preview:")
        print(df.head())
        
        print("\n✓ Excel processor working")
    else:
        print("Usage: python excel_processor.py <excel_file_path>")